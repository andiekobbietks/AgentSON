"""
AgentSON Excel Exporter - Export .agentson files to well-formatted Excel with charts and analytics.
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


def export_to_excel(input_path: str, output_path: str = None):
    """Export an .agentson file to Excel with charts and analytics."""
    
    # Load the .agentson file
    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create output path
    if output_path is None:
        output_path = str(Path(input_path).with_suffix('.xlsx'))
    
    # Create workbook
    wb = Workbook()
    
    # ============ SHEET 1: Overview ============
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    # Title
    ws_overview['A1'] = "AgentSON Session Overview"
    ws_overview['A1'].font = Font(name='Calibri', size=16, bold=True, color='2F5496')
    ws_overview.merge_cells('A1:E1')
    
    # Session metadata
    metadata = [
        ("Session ID", data.get('id', 'N/A')),
        ("Tool", data.get('tool', {}).get('name', 'N/A')),
        ("Agent", data.get('agent', {}).get('name', 'N/A')),
        ("Model", data.get('agent', {}).get('provider', 'N/A')),
        ("Started", data.get('started_at', 'N/A')),
        ("Ended", data.get('ended_at', 'N/A')),
        ("Total Entries", len(data.get('entries', []))),
        ("Total Tokens", data.get('metadata', {}).get('total_tokens', 'N/A')),
        ("Cost", data.get('metadata', {}).get('cost', 'N/A')),
    ]
    
    for i, (key, value) in enumerate(metadata, start=3):
        ws_overview[f'A{i}'] = key
        ws_overview[f'A{i}'].font = Font(bold=True)
        ws_overview[f'B{i}'] = str(value)
    
    # ============ SHEET 2: Entries Timeline ============
    ws_entries = wb.create_sheet("Entries Timeline")
    
    # Headers
    headers = ['Index', 'Type', 'Agent', 'Tool', 'Text/Code', 'Status', 'Timestamp']
    for col, header in enumerate(headers, 1):
        cell = ws_entries.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    
    # Data
    entries = data.get('entries', [])
    for idx, entry in enumerate(entries, 1):
        ws_entries.cell(row=idx+1, column=1, value=idx)
        ws_entries.cell(row=idx+1, column=2, value=entry.get('type', 'N/A'))
        ws_entries.cell(row=idx+1, column=3, value=entry.get('agent', 'N/A'))
        ws_entries.cell(row=idx+1, column=4, value=entry.get('tool', 'N/A'))
        
        # Truncate long text
        text = entry.get('text', entry.get('code', entry.get('query', '')))
        if len(str(text)) > 100:
            text = str(text)[:100] + '...'
        ws_entries.cell(row=idx+1, column=5, value=text)
        
        ws_entries.cell(row=idx+1, column=6, value=entry.get('status', 'N/A'))
        ws_entries.cell(row=idx+1, column=7, value=entry.get('timestamp', 'N/A'))
    
    # Auto-adjust column widths
    for column in ws_entries.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_entries.column_dimensions[column_letter].width = adjusted_width
    
    # ============ SHEET 3: Analytics ============
    ws_analytics = wb.create_sheet("Analytics")
    
    # Title
    ws_analytics['A1'] = "Session Analytics"
    ws_analytics['A1'].font = Font(name='Calibri', size=14, bold=True, color='2F5496')
    ws_analytics.merge_cells('A1:D1')
    
    # Entry type counts
    type_counts = {}
    for entry in entries:
        entry_type = entry.get('type', 'unknown')
        type_counts[entry_type] = type_counts.get(entry_type, 0) + 1
    
    ws_analytics['A3'] = "Entry Type Distribution"
    ws_analytics['A3'].font = Font(bold=True)
    
    ws_analytics['A4'] = "Type"
    ws_analytics['B4'] = "Count"
    ws_analytics['A4'].font = Font(bold=True)
    ws_analytics['B4'].font = Font(bold=True)
    
    for i, (entry_type, count) in enumerate(sorted(type_counts.items()), 5):
        ws_analytics[f'A{i}'] = entry_type
        ws_analytics[f'B{i}'] = count
    
    # Create pie chart for entry types
    chart = PieChart()
    chart.title = "Entry Type Distribution"
    chart.style = 10
    chart.width = 18
    chart.height = 12
    
    data_ref = Reference(ws_analytics, min_col=2, min_row=4, max_row=4+len(type_counts))
    cats = Reference(ws_analytics, min_col=1, min_row=5, max_row=4+len(type_counts))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    
    ws_analytics.add_chart(chart, "D3")
    
    # Tool usage counts
    tool_counts = {}
    for entry in entries:
        if entry.get('type') == 'action':
            tool = entry.get('tool', 'unknown')
            tool_counts[tool] = tool_counts.get(tool, 0) + 1
    
    tool_start_row = 5 + len(type_counts) + 2
    ws_analytics[f'A{tool_start_row}'] = "Tool Usage Distribution"
    ws_analytics[f'A{tool_start_row}'].font = Font(bold=True)
    
    ws_analytics[f'A{tool_start_row+1}'] = "Tool"
    ws_analytics[f'B{tool_start_row+1}'] = "Count"
    ws_analytics[f'A{tool_start_row+1}'].font = Font(bold=True)
    ws_analytics[f'B{tool_start_row+1}'].font = Font(bold=True)
    
    for i, (tool, count) in enumerate(sorted(tool_counts.items(), key=lambda x: x[1], reverse=True), tool_start_row+2):
        ws_analytics[f'A{i}'] = tool
        ws_analytics[f'B{i}'] = count
    
    # Create bar chart for tool usage
    if tool_counts:
        chart2 = BarChart()
        chart2.title = "Tool Usage Distribution"
        chart2.style = 10
        chart2.width = 18
        chart2.height = 12
        chart2.y_axis.title = "Count"
        
        data_ref2 = Reference(ws_analytics, min_col=2, min_row=tool_start_row+1, max_row=tool_start_row+1+len(tool_counts))
        cats2 = Reference(ws_analytics, min_col=1, min_row=tool_start_row+2, max_row=tool_start_row+1+len(tool_counts))
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.shape = 4
        
        ws_analytics.add_chart(chart2, f"D{tool_start_row}")
    
    # ============ SHEET 4: Errors & Issues ============
    ws_errors = wb.create_sheet("Errors & Issues")
    
    ws_errors['A1'] = "Errors and Issues"
    ws_errors['A1'].font = Font(name='Calibri', size=14, bold=True, color='C00000')
    ws_errors.merge_cells('A1:E1')
    
    error_headers = ['Index', 'Type', 'Tool', 'Error Message', 'Timestamp']
    for col, header in enumerate(error_headers, 1):
        cell = ws_errors.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    
    error_row = 4
    for idx, entry in enumerate(entries):
        if entry.get('status') == 'error' or entry.get('type') == 'error':
            ws_errors.cell(row=error_row, column=1, value=idx+1)
            ws_errors.cell(row=error_row, column=2, value=entry.get('type', 'N/A'))
            ws_errors.cell(row=error_row, column=3, value=entry.get('tool', 'N/A'))
            ws_errors.cell(row=error_row, column=4, value=entry.get('text', entry.get('output', 'N/A')))
            ws_errors.cell(row=error_row, column=5, value=entry.get('timestamp', 'N/A'))
            error_row += 1
    
    if error_row == 4:
        ws_errors.cell(row=4, column=1, value="No errors found")
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel exported to: {output_path}")
    return output_path


def export_all_to_excel(input_dir: str = "examples", output_dir: str = "exports"):
    """Export all .agentson files in a directory to Excel."""
    
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    agentson_files = list(input_path.glob("*.agentson"))
    
    if not agentson_files:
        print(f"No .agentson files found in {input_dir}")
        return
    
    print(f"Found {len(agentson_files)} .agentson files")
    
    for agentson_file in agentson_files:
        try:
            output_file = output_path / f"{agentson_file.stem}.xlsx"
            export_to_excel(str(agentson_file), str(output_file))
        except Exception as e:
            print(f"Error exporting {agentson_file.name}: {e}")
    
    print(f"\nExported {len(agentson_files)} files to {output_dir}")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Export AgentSON to Excel")
    parser.add_argument("input", help="Input .agentson file or directory")
    parser.add_argument("--output", help="Output .xlsx file or directory")
    parser.add_argument("--all", action="store_true", help="Export all .agentson files in directory")
    args = parser.parse_args()
    
    if args.all:
        export_all_to_excel(args.input, args.output or "exports")
    else:
        export_to_excel(args.input, args.output)
